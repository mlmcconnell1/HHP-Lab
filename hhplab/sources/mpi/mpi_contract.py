"""Source contract for MPI unauthorized immigrant topline workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Final

MPI_SOURCE_ID: Final = "mpi_unauthorized_immigrants"
MPI_PROVIDER: Final = "migration_policy_institute"
MPI_PRODUCT: Final = "unauthorized_immigrant_topline_estimates"
MPI_PUBLICATION_YEAR: Final = 2025
MPI_ESTIMATE_YEAR: Final = 2023
MPI_ESTIMATE_PERIOD: Final = "mid_2023"
MPI_WORKBOOK_GLOB: Final = "MPI-2023_Unauthorized_Profiles-State-County-Topline_Estimates*.xlsx"

MPI_SOURCE_PAGE: Final = (
    "https://www.migrationpolicy.org/research/unauthorized-immigrants-us-2025-fact-sheet"
)
MPI_SOURCE_URL: Final = (
    "https://www.migrationpolicy.org/sites/default/files/content-media/2025/10/"
    "MPI-2023_Unauthorized_Profiles-State-County-Topline_Estimates-FINAL.xlsx"
)

MPI_STATE_SHEET: Final = "U.S. and States"
MPI_COUNTY_SHEET: Final = "U.S. and Counties"
MPI_REQUIRED_SHEETS: Final = (MPI_STATE_SHEET, MPI_COUNTY_SHEET)
MPI_TITLE_ROW: Final = 2
MPI_HEADER_ROW: Final = 3
MPI_FIRST_DATA_ROW: Final = 4

MPI_STATE_HEADERS: Final = (
    "State",
    "Number of Unauthorized Immigrants",
    "State Share of the Total Unauthorized Immigrant Population",
)
MPI_COUNTY_HEADERS: Final = (
    "State",
    "County",
    "Number of Unauthorized Immigrants",
    "County Share of the Total Unauthorized Immigrant Population",
)

MPI_RAW_STATE_COLUMNS: Final = (
    "source_sheet",
    "source_row",
    "state_name",
    "unauthorized_immigrant_population",
    "unauthorized_immigrant_share_of_us_total",
    "year",
    "estimate_period",
    "geo_type",
    "state_fips",
    "is_us_total",
)
MPI_RAW_COUNTY_COLUMNS: Final = (
    "source_sheet",
    "source_row",
    "state_name",
    "county_label",
    "unauthorized_immigrant_population",
    "unauthorized_immigrant_share_of_us_total",
    "year",
    "estimate_period",
    "geo_type",
    "county_fips",
    "state_fips",
    "is_us_total",
    "is_county_equivalent",
    "exclusion_reason",
)
MPI_CURATED_COUNTY_COLUMNS: Final = (
    "geo_type",
    "geo_id",
    "county_fips",
    "state_fips",
    "state_name",
    "county_label",
    "year",
    "estimate_period",
    "unauthorized_immigrant_population",
    "unauthorized_immigrant_share_of_us_total",
    "source_id",
    "provider",
    "product",
    "data_source",
    "source_url",
    "source_sheet",
    "source_row",
    "raw_sha256",
    "ingested_at",
)
MPI_MEASURE_COLUMNS: Final = (
    "unauthorized_immigrant_population",
    "unauthorized_immigrant_share_of_us_total",
)
MPI_REQUIRED_CURATED_COLUMNS: Final = ("county_fips", "year")

MPI_SOURCE_CITATION: Final = (
    "Migration Policy Institute analysis of U.S. Census Bureau 2019-23 ACS and "
    "2023, along with 2008, SIPP data, weighted to 2023 unauthorized immigrant "
    "population estimates provided by Jennifer Van Hook of The Pennsylvania State "
    "University."
)
MPI_METHODOLOGY_NOTE: Final = (
    "MPI assigns legal status to noncitizens in ACS records using SIPP-informed "
    "characteristics and benchmarks estimates to control totals. Estimates carry "
    "survey sampling and coverage error and should be interpreted as mid-2023 "
    "point-in-time estimates, not annual flow counts."
)

MPI_ALLOWED_COUNTY_EQUIVALENT_LABELS: Final = (
    "County",
    "Parish",
    "Borough",
    "Census Area",
    "Municipality",
    "City",
)
MPI_EXCLUDED_COUNTY_ROW_PATTERNS: Final = (
    "United States",
    "Counties",
    "Parishes",
    "MSA",
)
MPI_GEOGRAPHY_RULES: Final = (
    "State sheet rows are source totals and are not county-native panel rows.",
    "County sheet row 4 is the U.S. total and must be excluded from county output.",
    "Rows naming one county-equivalent, independent city, Alaska municipality, "
    "borough, census area, or Louisiana parish may be curated only after resolving "
    "a unique county_fips within the stated state.",
    "Rows naming multiple counties, multiple parishes, partial counties, or MSAs "
    "must be retained only in raw-normalized diagnostics or excluded with an "
    "exclusion_reason until an explicit source-specific crosswalk exists.",
    "County labels are resolved using the state column plus the state name embedded "
    "in the county label; county names alone are never globally unique.",
)


@dataclass(frozen=True)
class MpiWorkbookContract:
    """Declarative layout and schema contract for an MPI workbook release."""

    source_id: str = MPI_SOURCE_ID
    workbook_glob: str = MPI_WORKBOOK_GLOB
    required_sheets: tuple[str, ...] = MPI_REQUIRED_SHEETS
    header_row: int = MPI_HEADER_ROW
    first_data_row: int = MPI_FIRST_DATA_ROW
    state_headers: tuple[str, ...] = MPI_STATE_HEADERS
    county_headers: tuple[str, ...] = MPI_COUNTY_HEADERS
    raw_state_columns: tuple[str, ...] = MPI_RAW_STATE_COLUMNS
    raw_county_columns: tuple[str, ...] = MPI_RAW_COUNTY_COLUMNS
    curated_county_columns: tuple[str, ...] = MPI_CURATED_COUNTY_COLUMNS
    measure_columns: tuple[str, ...] = MPI_MEASURE_COLUMNS
    geography_rules: tuple[str, ...] = MPI_GEOGRAPHY_RULES


MPI_WORKBOOK_CONTRACT: Final = MpiWorkbookContract()


def validate_mpi_workbook_contract(path: Path | str) -> MpiWorkbookContract:
    """Validate the workbook layout needed by the MPI ingest implementation."""
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(
            f"MPI workbook not found: {workbook_path}. Expected a file matching "
            f"{MPI_WORKBOOK_GLOB} under data/raw/mpi; stage the workbook from "
            f"{MPI_SOURCE_URL}."
        )
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(
            f"Unsupported MPI workbook type '{workbook_path.suffix}'. Expected an "
            f".xlsx file matching {MPI_WORKBOOK_GLOB}."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is present in project env
        raise RuntimeError(
            "openpyxl is required to inspect MPI XLSX workbooks; install the "
            "project dependencies with `uv sync --extra dev`."
        ) from exc

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing_sheets = [sheet for sheet in MPI_REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(
            f"Unsupported MPI workbook layout: missing sheets {missing_sheets}. "
            f"Expected sheets {list(MPI_REQUIRED_SHEETS)} with headers on row "
            f"{MPI_HEADER_ROW}."
        )

    _validate_headers(
        workbook[MPI_STATE_SHEET],
        expected=MPI_STATE_HEADERS,
        sheet_name=MPI_STATE_SHEET,
    )
    _validate_headers(
        workbook[MPI_COUNTY_SHEET],
        expected=MPI_COUNTY_HEADERS,
        sheet_name=MPI_COUNTY_SHEET,
    )
    return MPI_WORKBOOK_CONTRACT


def _validate_headers(worksheet, *, expected: tuple[str, ...], sheet_name: str) -> None:
    actual = tuple(
        worksheet.cell(MPI_HEADER_ROW, column_index).value
        for column_index in range(1, len(expected) + 1)
    )
    if actual != expected:
        raise ValueError(
            f"Unsupported MPI workbook layout on sheet '{sheet_name}': expected "
            f"headers {list(expected)} on row {MPI_HEADER_ROW}, found {list(actual)}. "
            "Update the MPI source contract before ingesting this workbook."
        )
